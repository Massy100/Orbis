from datetime import time
from openpyxl import load_workbook


VALID_MARKS = {"x", "X", "1", "si", "sí", "SI", "SÍ", "true", "TRUE"}


def parse_time_range(value):
    """
    Transform "07:00 - 10:00" into two time objects.
    """
    start_str, end_str = str(value).split("-")
    start_str = start_str.strip()
    end_str = end_str.strip()

    start_hour, start_minute = map(int, start_str.split(":"))
    end_hour, end_minute = map(int, end_str.split(":"))

    return time(start_hour, start_minute), time(end_hour, end_minute)


def is_marked(cell):
    """
    Detect if the cell is marked as selected based on its value or background color.

    accept:
    - cell value in VALID_MARKS
    - cell with background color
    """

    if cell.value is not None:
        text = str(cell.value).strip()

        if text in VALID_MARKS:
            return True

    fill = cell.fill

    if not fill or fill.fill_type is None:
        return False

    fg_color = fill.fgColor

    if fg_color is None:
        return False

    if fg_color.type == "rgb":
        rgb = fg_color.rgb
        return rgb not in (None, "00000000", "FFFFFFFF")

    if fg_color.type in ("indexed", "theme"):
        return True

    return False


def parse_teacher_schedule_excel(file, teacher_code):
    """
    Read the Excel and return the marked schedules for a teacher.

    Structure:
    - Row 2: days.
    - Row 3: time ranges.
    - Column B: Code.
    - Column C: Teacher.
    - From column D onwards: schedules.
    """

    wb = load_workbook(file, data_only=True)
    ws = wb.active

    teacher_code = str(teacher_code).strip()

    schedule_columns = {}

    current_day = None

    for col in range(4, ws.max_column + 1):
        day_cell = ws.cell(row=2, column=col).value
        time_cell = ws.cell(row=3, column=col).value

        if day_cell is not None and str(day_cell).strip():
            current_day = str(day_cell).strip()

        if current_day and time_cell:
            start_time, end_time = parse_time_range(time_cell)

            schedule_columns[col] = {
                "day": current_day,
                "starttime": start_time,
                "endtime": end_time,
            }

    found_teacher_name = None
    selected_schedules = []

    for row in range(4, ws.max_row + 1):
        code_cell = ws.cell(row=row, column=2).value

        if code_cell is None:
            continue

        code = str(code_cell).strip()

        if code != teacher_code:
            continue

        found_teacher_name = ws.cell(row=row, column=3).value

        for col, schedule_data in schedule_columns.items():
            cell = ws.cell(row=row, column=col)

            if is_marked(cell):
                selected_schedules.append(schedule_data)

        break

    if found_teacher_name is None:
        raise ValueError(f"No se encontró el docente con código {teacher_code} en el archivo.")

    return {
        "teacher_code": teacher_code,
        "teacher_name": found_teacher_name,
        "schedules": selected_schedules,
    }